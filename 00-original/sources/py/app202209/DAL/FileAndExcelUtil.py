class FileAndExcelUtil:
    @staticmethod
    def _CopyData(fromSheet, toSheet, rowOffset):
        rowI = 0
        for row in range(fromSheet.max_row):
            for column in range(fromSheet.max_column):
                columnletter=chr(column+1+65-1)#cell.get_column_letter(column+1)
                cellName = "%s%s"%(columnletter,row+1)
                if(fromSheet[cellName].value != None):
                    toSheet[cellName].value = fromSheet[cellName].value    
                    rowI = row + 1                
                else:
                    pass
        for I in range(500+rowOffset,rowI,-1):
            toSheet.delete_rows(I, 1)

    @staticmethod
    def _MakeFormat(datPathAndFile, srcFile, formatFile, toFile):
        from openpyxl import load_workbook, cell
        import os
        fullSrcFile = os.path.join(datPathAndFile,f'{srcFile}.xlsx')
        fullFormatFile = os.path.join(datPathAndFile,f'{formatFile}.xlsx')
        fullToFile = os.path.join(datPathAndFile,f'{toFile}.xlsx')
        rowOffset = 0 #if int(formatFile[-1]) == 1 else 1

        srcBk = load_workbook(fullSrcFile)
        srcSheets = srcBk.sheetnames
        #print(srcSheets)

        fmtBk = load_workbook(fullFormatFile)
        fmtSheets = fmtBk.sheetnames
        #print(fmtSheets)

        for I in range(len(fmtSheets)-1,len(srcSheets)-1,-1):
            fmtBk.remove_sheet(fmtBk[fmtSheets[I]])
        #print(fmtBk.sheetnames)


        for I in range(len(srcSheets)):
            fmtBk[fmtSheets[I]].title = srcSheets[I]
            srcSheetxl = srcBk[srcSheets[I]]
            fmtSheetxl = fmtBk[srcSheets[I]]
            FileAndExcelUtil._CopyData(srcSheetxl, fmtSheetxl, rowOffset)

        #print(fmtBk.sheetnames)
        
        fmtBk.save(fullToFile)
    @staticmethod
    def RemoveFile(file):
        import os
        print(file)
        if input("Are you sure to delete(y/n)?") != "y":
            return
        os.remove(file)