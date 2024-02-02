import os
import pyexcel as pyxl
from datetime import datetime
from .MkscoreUtil import MkscoreUtil

class FormatXlBook:
    @staticmethod
    def CopyData(fromSheet, toSheet, rowOffset):
        rowI = 0
        for row in range(fromSheet.max_row):
            for column in range(fromSheet.max_column):
                columnletter=chr(column+1+65-1)#cell.get_column_letter(column+1)
                cellName = "%s%s"%(columnletter,row+1)
                if(fromSheet[cellName].value != None):
                    toSheet[cellName].value = fromSheet[cellName].value    
                    rowI = row + 1                
                else:
                    pass
        for I in range(199+rowOffset,rowI,-1):
            toSheet.delete_rows(I, 1)
        
    @staticmethod
    def MakeFormat(datPathAndFile, srcFile, formatFile, toFile):
        from openpyxl import load_workbook, cell
        fullSrcFile = os.path.join(datPathAndFile,f'dest/{srcFile}.xlsx')
        fullFormatFile = os.path.join(datPathAndFile,f'fmt/{formatFile}.xlsx')
        fullToFile = os.path.join(datPathAndFile,f'mail/{toFile}.xlsx')
        rowOffset = 0 if int(formatFile[-1]) == 1 else 1

        srcBk = load_workbook(fullSrcFile)
        srcSheets = srcBk.sheetnames
        srcSheets.remove('Consolidated')
        #print(srcSheets)

        fmtBk = load_workbook(fullFormatFile)
        fmtSheets = fmtBk.sheetnames
        #print(fmtSheets)

        for I in range(len(fmtSheets)-1,len(srcSheets)-1,-1):
            fmtBk.remove_sheet(fmtBk[fmtSheets[I]])
        #print(fmtBk.sheetnames)


        for I in range(len(srcSheets)):
            fmtBk[fmtSheets[I]].title = srcSheets[I]
            srcSheetxl = srcBk[srcSheets[I]]
            fmtSheetxl = fmtBk[srcSheets[I]]
            FormatXlBook.CopyData(srcSheetxl, fmtSheetxl, rowOffset)

        #print(fmtBk.sheetnames)
        
        fmtBk.save(fullToFile)

class ManyTest:
    @staticmethod
    def readTemplateBool(singleData):
        #Data Input
        testName,               md_templates,               md_soruces,                destFileName,            metaData,               templateBk,                 isCode          = singleData['testName'], singleData['md_templates'],singleData['md_soruces'],singleData['destFileName'],singleData['metaData'],singleData['templateBk'],singleData['isCode']        
        #???print('Reading Template Book...')        
        #Processing
        templateBk = MkscoreUtil.readBkByCols(
                os.path.join(md_templates['path'],md_templates['file']), 
                md_templates['sheets'], 
                md_templates['cols'], 
                1)

        #???print('Reading Source Book...')        
        #   **NEW**    iterate here : src(es)
        pref_sources = md_soruces
        pref_source = pref_sources
        
     
        metaData = {'xlBk':[], 'testName':[],'part' :[], 'colDef':[], 'isCodeRight':[]}
        for i in range(0,len(pref_sources)):
            pref_source = pref_sources[i] 
            #print(pref_source)
            if i == 0:
                metaData['date'] = pref_source['dsname'][0] # one time
                dt = datetime.strptime(metaData['date'], '%d/%m/%Y')
                metaData['date_d'] = dt.strftime('%d-%b')
                metaData['day_name'] = dt.strftime('%a')
                metaData['day_name_full'] = dt.strftime('%A')
            
            metaData['testName'].append(pref_source['dsname'][2])
            metaData['part'].append(pref_source['dsname'][1])
            metaData['isCodeRight'].append(not pref_source['isMCQ'])
            #First source book
            sourceBk =  MkscoreUtil.readBkByCols(
                    os.path.join(pref_source['path'],pref_source['file']), 
                    [], 
                    pref_source['cols'], 
                    int(pref_source['headerRow']),True)
            metaData['colDef'].append([colName for colName in sourceBk[0]['data'][0]])
            metaData['xlBk'].append(sourceBk)           

        
        dtNow = datetime.now().strftime('%Y%m%d%H%M%S') 
        metaData['part'] = [e[-1] for e in metaData['testName']]       
        destFileName = '{0}-{1}-{2}-{3}-Scores-{4}.xlsx' .format (
            md_templates['testLinkName'],
            metaData['day_name'],
            metaData['date_d'],
            metaData['testName'][0][:-1].strip() + '_'.join(metaData['part']),
            dtNow
            )     
        testName = metaData['testName']
        
        #Data Output
        singleData['testName'], singleData['md_templates'],singleData['md_soruces'],singleData['destFileName'],singleData['metaData'],singleData['templateBk'],singleData['isCode'] = testName, md_templates, md_soruces, destFileName, metaData,templateBk, isCode
        
    @staticmethod
    def readStudentEachTestScore(singleData):
        #Data Input
        testName,               md_templates,               md_soruces,                destFileName,            metaData,               templateBk,                 isCode          = singleData['testName'], singleData['md_templates'],singleData['md_soruces'],singleData['destFileName'],singleData['metaData'],singleData['templateBk'],singleData['isCode']
        #Processing
        '''
            input data:
            (1)
                    []  {}                  []       {}        primitive 
            templateBk Sheet                Rows    Row    key:value
                     [  {name:Batch 1, data:[       {       col:value ....} ....]  } ]
 

            (2)
                    []  {}                  []       {}        primitive 
                srcBk Sheet                Rows    Row    key:value
                     [  {name:Batch 1, data:[       {       col:value ....}} ....]   ]                     
                    ...
                binds with testName
            ==>
            
            (1)   
                []  {}                  []       {}        primitive 
                templateBk Sheet                Rows    Row    key:value
                        [  {name:Batch 1, data:[       {       col:value ....
                                                                (+)Recent: {score:20,attendance:xx,remarks:xx}....
                                                        }                ....
                                                ], 
                        

                            (+)Recent: {'Attendance':{'PRESENT':103, ...},'Remarks:{'ScoreNot':121, ....}} ,
                            ....
                            (+)'sheetMatrix' :
                                [
                                    ['S. No.','Name of Student','Dept', 'Score', 'Remarks', 'Attendance'],
                                    [101,       Prabhu,          ECE,    20,     ScoreCant,   PRESENT]...
                                ] 
                            (+) Recent_sum : {'ABSENT':0, 'PRESENT':0, 'Good job. Keep it up!':0, 'Attended test in a Genuine manner':0}
                            }  
                        ] 
                        In short, sheet we added 'sheetMatrix' key
                            which holds the printable data for the sheet
                            [
                                ['S. No.','Name of Student','Dept', 'Score', 'Remarks', 'Attendance'],
                                [101,       Prabhu,          ECE,    20,     ScoreCant,   PRESENT]...
                            ]

            (3)
            bookdict 
                !dict {testname: printable matrix of test data}
                !testname will be sheet name
            sumTestMatrices 
                !list of list(each test) of conosolidated records
            sumMatrix           
                !printable matrix of "consolidated" data 
                !single list of "consolidated" records of all tests
        '''
        
        
        testNameParts = metaData['part']
        testNames = testName
        testCount = len(testNames)

        #Printable Each Sheet 
        partHeadingsDef = lambda : ['',''] #Student Detail Part Heading ,''
        capHeadingsDef = lambda : ['S. No.','Name of Student'] #Student Detail Caption Heading ,'Dept'
        sheetMatrixDef = lambda : [ partHeadingsDef(), capHeadingsDef() ]
        #End Printable Each Sheet 

        #Footer Data
        footDataDef = lambda  : {'ABSENT':0, 'PRESENT':0, 'Good job. Keep it up!':0, 'Attended test in a Genuine manner':0}
        sumTestMatrixDef = lambda : [
                    ['', 'Batch',	'Total',	'Attended',	'Absent',	'Good Scorers',	'Genuine Attender', 	'Just Attended'],
                    ['','Total',0,0,0,0,0,0],
                    ['','','','','','','','']
                ]
        sumTestMatrices = [] #list of test wise 'consolidated' data
        sumMatrix = []   #list of lists  transformed to list ie printable consolidated data
        #End Footer Data
        
        matchCol = md_templates['MatchHeading']    
        
        bookdict = {}

        for testI in range(len(testNames)):
            testName = testNames[testI]   
            testNamePart = testNameParts[testI]        
              
            xlSheet = metaData['xlBk'][testI][0]
            colDef = metaData['colDef'][testI]
            isCodeRight = metaData['isCodeRight'][testI]
            #Footer Data
            sheetNum = 0  
            sumTestMatrix = sumTestMatrixDef()
            sumTestMatrices.append(sumTestMatrix)
            #End Footer Data
            for sheet in templateBk:  #sheet -> dict where key(s) 'data'  
                #Printable Each Sheet                 
                sMi = 0
                sheetMatrix = MkscoreUtil.DictSetGetKey(sheet, 'sheetMatrix',sheetMatrixDef()) #Printable Data
                bookdict[sheet['name']] = sheetMatrix
                
                sMPartHeadings = sheetMatrix[sMi]; sMi+=1
                sMCapHeadings = sheetMatrix[sMi]; sMi+=1

                #(+)Scores Part Heading
                sMPartHeadings.extend(['Part ' + testNamePart,'']) #,''              
                #(+)Scores Caption Heading
                sMCapHeadings.extend([colDef[-3],colDef[-1]]) #colDef[-3:]
                #End Printable Each Sheet 

                #Footer Data
                shTestName = {'Attendance':{}, 'Remarks': {}}                
                sheet[testName] = shTestName  #sheet keys: 'data', testName (+)               
                shTestNameSum = footDataDef()
                sheet[testName + '_sum'] = shTestNameSum
                #End Footer Data               

                shData = sheet['data']
                for row in shData:       
                    matchData = row[matchCol] #ID Column Data

                    #Printable Each Sheet 
                    if len(sheetMatrix) > sMi:
                        sMRow = sheetMatrix[sMi]; sMi+=1 #If there
                    else:
                        sMRow = [row['S. No.'],row['Name of Student']] #Student Detail Data #,row['Dept']
                        sheetMatrix.append(sMRow); sMi+=1
                    #End Printable Each Sheet 

                    #READ & MUTE the SOURCE attendance, remarks, scores 
                    testData = MkscoreUtil.ReadDataFromBookv2(xlSheet, matchData, matchCol, colDef, isCodeRight)
                    row[testName] = testData
                    
                    #Printable Each Sheet 
                    sMRow.extend([testData['score'],testData['remarks']]) # (+)Scores Data (of Each Part) ,testData['attendance']
                    #End Printable Each Sheet 

                    #Footer Data
                    MkscoreUtil.setCountInDict(shTestName['Attendance'],testData['attendance']) #for example 'PRESENT' will become key
                    MkscoreUtil.setCountInDict(shTestName['Remarks'],testData['remarks'])#for example 'Scores cant be displayed' will become key
                    #End Footer Data
                #end for row    
                    
                
                #Footer Data
                dbOneABSENT = shTestName['Attendance'].get('ABSENT',0)
                dbOnePRESENT =  shTestName['Attendance'].get('PRESENT',0)
                dbOneGoodScorers =  shTestName['Remarks'].get('Good job. Keep it up!',0) 
                dbOneGenuineAttender = shTestName['Remarks'].get('Attended test in a Genuine manner',0)

                shTestNameSum['ABSENT'] = dbOneABSENT
                shTestNameSum['PRESENT'] = dbOnePRESENT
                shTestNameSum['Good job. Keep it up!'] = dbOneGoodScorers
                shTestNameSum['Attended test in a Genuine manner'] = dbOneGenuineAttender
                #sumTestMatrix
                sTMI = -2                    
                dbOneTotal = dbOneABSENT + dbOnePRESENT                
                dbOneJustAttended = dbOnePRESENT - dbOneGoodScorers - dbOneGenuineAttender
                sumTestMatrix.insert(sTMI,['', md_templates['sheets'][sheetNum],	
                    dbOneTotal,	
                    dbOnePRESENT, dbOneABSENT,	
                    dbOneGoodScorers, dbOneGenuineAttender, dbOneJustAttended])

                if testCount > 1: sumTestMatrix[0][1] = 'Part ' + testNamePart #part heading is not for single test
                sumTestMatrix[sTMI][2] += dbOneTotal
                sumTestMatrix[sTMI][3] += dbOnePRESENT
                sumTestMatrix[sTMI][4] += dbOneABSENT
                sumTestMatrix[sTMI][5] += dbOneGoodScorers
                sumTestMatrix[sTMI][6] += dbOneGenuineAttender
                sumTestMatrix[sTMI][7] += dbOneJustAttended
                #End Footer Data

                sheetNum += 1
            #end for sheet

            #Footer Data
            sumMatrix.extend(sumTestMatrix)
            #End Footer Data
        #end for testI    

        #Writable Book adds the 'Consolidated'
        if testCount == 1:                          #part heading is not for single test
            for e in bookdict: bookdict[e].pop(0)   #part heading is not for single test
        bookdict['Consolidated'] = sumMatrix
            

        #Writeable XL Book with Sheets and save as excel book file    
        fileName = destFileName
        srcPath = md_soruces[0]['path']
        jobPath = srcPath[0:-4]
        destPath = os.path.join(jobPath,'dest')
        dest_file_name = os.path.join(destPath, fileName)      
        
        #save
        pyxl.save_book_as(bookdict=bookdict,  dest_file_name = dest_file_name)


        #update as job has DONE
        
        for eSour in md_soruces:
            srcFilePath = os.path.join(srcPath,eSour['file'])
            os.rename(srcFilePath, srcFilePath[0:-4] + '-done.csv')
    
        

        #???print('!!!Thank you!!!')   
        #Data Output
        singleData['testName'], singleData['md_templates'],singleData['md_soruces'],singleData['destFileName'],singleData['metaData'],singleData['templateBk'],singleData['isCode'] = testName, md_templates, md_soruces, destFileName, metaData,templateBk, isCode
        #return these data from here
        retData = {
            'jobPath' : jobPath,
            'destPath' : destPath,
            'destFile' : fileName,
            'destFileOnly' : fileName.split(".")[0],
            'destPathFile' : dest_file_name,
            'mailPath' : os.path.join(jobPath,'mail'),
            'mailFile' : "-".join(fileName.split(".")[0].split("-")[:-1]) + '.xlsx',
            'mailFileOnly' : "-".join(fileName.split(".")[0].split("-")[:-1])
        } 
        #Tested code: '_'.join('a-b-c.xlsx'.split('.')[0].split('-')[:-1]) + '.xlsx'
        retData['mailFilePath'] = os.path.join(retData['mailPath'],retData['mailFile'])
        FormatXlBook.MakeFormat(jobPath,retData['destFileOnly'], f"fmt{len(singleData['md_soruces'])}", retData['mailFileOnly'])
        #MkscoreUtil.openWorkBook(md_templates['formatFile'])
        MkscoreUtil.openWorkBook(dest_file_name)
        MkscoreUtil.openWorkBook(retData['mailFilePath'])

        return retData