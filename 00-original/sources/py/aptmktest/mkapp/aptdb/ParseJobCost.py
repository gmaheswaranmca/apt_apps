import os
import pyexcel as pyxl
from datetime import datetime
import sys


class ParseJobCost:
    @staticmethod
    def ReadJobFile(jobPath, jobFile, fromDate, toDate):
        from openpyxl import load_workbook, cell
        jobSrcFile = os.path.join(jobPath,f'{jobFile}.xlsx')

        srcBk = load_workbook(jobSrcFile)
        srcSheets = srcBk.sheetnames
        srcSheets.remove('Dates')
        srcSheets.remove('JobType')
        #print(srcSheets)
        proData = {}
        linkData = {}
        #{'10-Oct-2019':{'odate':'','ddate':'','day':'','tests':[{'testName','job','cost'}],printables:''}}

        for I in range(len(srcSheets)): #Iterate Sheets testName
            testName = srcSheets[I]
            theSheet = srcBk[testName]
            #print(testName)
            linkData[testName] = linkData.get(testName,{'SP':[0,0],'SP-C':[0,0],'MP':[0,0],'MP-C':[0,0],'NID':[0,0],'AID':[0,0],'Total':[0,0]}) 
            for row in range(1,theSheet.max_row):#Iterate Rows
                '''Read Data date,job '''
                colDt = "B"
                cellDt = "%s%s"%(colDt,row+1)
                if(theSheet[cellDt].value != None):
                    lineJobs = ParseJobCost.ReadLineJob(theSheet,row+1)
                    for lineJob in lineJobs:
                        odate = lineJob['date']
                        if not(fromDate <= odate and odate <= toDate):
                            continue
                        ddate = lineJob['ddate']
                        day  = lineJob['day']
                        defVal = {'odate':odate,'ddate':ddate,'day':day,'tests':[]}
                        proData[odate] = proData.get(odate,defVal)
                        proData[odate]['tests'].append({'testName':testName,'job':lineJob['job'],'jobDesc':lineJob['jobDesc'],'jobCost':lineJob['jobCost']})
                        linkData[testName][lineJob['job']][0] += 1
                        linkData[testName][lineJob['job']][1] += lineJob['jobCost']
                        linkData[testName]['Total'][0] += 1
                        linkData[testName]['Total'][1] += lineJob['jobCost']
                        #print(lineJob['date'].strftime('%Y-%m-%d'))
                        #print(lineJob)
                    '''if row==5:
                        break'''
        

        pdata = []
        
        I=0
        for e in proData:
            I+=1
            #print(e,proData[e])
            desc = ''
            J = 0
            jobCost = 0
            for eJob in proData[e]['tests']:
                J += 1
                desc = desc + (', ' if len(desc)>0 else '')
                desc = desc + '%d-%s[%s]'%(J,eJob['testName'],eJob['jobDesc'])
                jobCost += eJob['jobCost']
            pdata.append([e, I, proData[e]['ddate'], proData[e]['day'], desc, jobCost, J])
        pdata = sorted(pdata, key=lambda x:x[0])
        I=0
        for e in pdata:
             I+=1
             e[1] = I

        lData = []
        I=0
        for e in linkData:
            if linkData[e]['Total'][0] <= 0:
                continue
            I+=1
            lData.append([e,I,e,
                linkData[e]['SP'][0],linkData[e]['SP'][1],
                linkData[e]['SP-C'][0],linkData[e]['SP-C'][1],
                linkData[e]['MP'][0],linkData[e]['MP'][1],
                linkData[e]['MP-C'][0],linkData[e]['MP-C'][1],
                linkData[e]['NID'][0],linkData[e]['NID'][1],
                linkData[e]['AID'][0],linkData[e]['AID'][1],
                linkData[e]['Total'][0],linkData[e]['Total'][1]
                ])
        lData = sorted(lData, key=lambda x:x[0])
        I=0
        for e in lData:
             I+=1
             e[1] = I

        '''for e in lData:
            print(e)'''

        fullFormatFile = os.path.join(jobPath,'jobcostfmt.xlsx')
        
        fullToFile = os.path.join(jobPath,'jobcost/%s_to_%s.xlsx'%(fromDate.strftime("%Y-%m%b-%d%a"),toDate.strftime("%Y-%m%b-%d%a")))
        fmtBk = load_workbook(fullFormatFile)
        fmtSheets = fmtBk.sheetnames
        for SheetNo in range(2):
            wrSheet = fmtBk[fmtSheets[SheetNo]]
            vData = [pdata,lData][SheetNo]
            rowI = 0
            for row in range(len(vData)):
                for column in range(wrSheet.max_column):
                    columnletter=chr(column+1+65-1) #cell.get_column_letter(column+1)
                    cellName = "%s%s"%(columnletter,row+2)
                    wrSheet[cellName].value = vData[row][column+1] 
        #print(fmtBk.sheetnames)
        
        fmtBk.save(fullToFile)

    @staticmethod
    def ReadLineJob(theSheet, rowNum):
        colDt = "B"
        colJob = "G"
        cellDt = "%s%s"%(colDt,rowNum)
        cellJob = "%s%s"%(colJob,rowNum)
        theDate = theSheet[cellDt].value
        theDDate = theDate.strftime('%d-%b-%Y')
        theDay = theDate.strftime('%a')
        job = theSheet[cellJob].value.split(",")
        jobs = []
        jobDescType = {'SP':['Test',125],
            'SP-C':['Test-Coding',150],
            'MP':['Multipart Test',250],
            'MP-C':['Multipart Test-1 Coding',275],
            'NID':['New IDs',125],
            'AID':['Add IDs',125]
        }
        for e in job:
            e = e.strip()
            jobDesc = jobDescType[e][0]
            jobCost = jobDescType[e][1]
            jobs.append({'date': theDate, 'ddate':theDDate,'day':theDay,
                'job': e, 'jobDesc':jobDesc,'jobCost':jobCost})
        return jobs

    @staticmethod
    def CopyData(fromSheet, toSheet, rowOffset):
        rowI = 0
        for row in range(fromSheet.max_row):
            for column in range(fromSheet.max_column):
                columnletter=chr(column+1+65-1)#cell.get_column_letter(column+1)
                cellName = "%s%s"%(columnletter,row+1)
                if(fromSheet[cellName].value != None):
                    toSheet[cellName].value = fromSheet[cellName].value    
                    rowI = row + 1                
                else:
                    pass
        for I in range(199+rowOffset,rowI,-1):
            toSheet.delete_rows(I, 1)
        
    @staticmethod
    def MakeFormat(datPathAndFile, srcFile, formatFile, toFile):
        from openpyxl import load_workbook, cell
        fullSrcFile = os.path.join(datPathAndFile,f'dest/{srcFile}.xlsx')
        fullFormatFile = os.path.join(datPathAndFile,f'fmt/{formatFile}.xlsx')
        fullToFile = os.path.join(datPathAndFile,f'mail/{toFile}.xlsx')
        rowOffset = 0 if int(formatFile[-1]) == 1 else 1

        srcBk = load_workbook(fullSrcFile)
        srcSheets = srcBk.sheetnames
        srcSheets.remove('Consolidated')
        #print(srcSheets)

        fmtBk = load_workbook(fullFormatFile)
        fmtSheets = fmtBk.sheetnames
        #print(fmtSheets)

        for I in range(len(fmtSheets)-1,len(srcSheets)-1,-1):
            fmtBk.remove_sheet(fmtBk[fmtSheets[I]])
        #print(fmtBk.sheetnames)


        for I in range(len(srcSheets)):
            fmtBk[fmtSheets[I]].title = srcSheets[I]
            srcSheetxl = srcBk[srcSheets[I]]
            fmtSheetxl = fmtBk[srcSheets[I]]
            FormatXlBook.CopyData(srcSheetxl, fmtSheetxl, rowOffset)

        #print(fmtBk.sheetnames)
        
        fmtBk.save(fullToFile)
    @staticmethod
    def Do():        
        jobPath, jobFile = r"F:\v2m\techLang\assignment\2020-07-08-Apt-BulkQuestions\OLT\OLTAssignment\scoresfiles\app_master","job-log"
        if len(sys.argv) == 3:
            fromDate, toDate = sys.argv[1], sys.argv[2]  
            fromDate = datetime.strptime(fromDate, '%Y%b%d')
            toDate = datetime.strptime(toDate, '%Y%b%d')
            ParseJobCost.ReadJobFile(jobPath, jobFile, fromDate, toDate)

    
ParseJobCost.Do()        